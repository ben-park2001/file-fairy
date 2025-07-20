"""Extractors for data files (CSV, Excel, JSON)."""

import json
import csv
from pathlib import Path
from .base import BaseExtractor


class CSVExtractor(BaseExtractor):
    """Extractor for CSV files."""
    
    def extract(self, file_path: Path) -> str:
        """Extract text from CSV files."""
        try:
            content_rows = []
            with open(file_path, 'r', encoding='utf-8') as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    content_rows.append('\t'.join(row))
            return '\n'.join(content_rows)
        except Exception as e:
            return self._handle_extraction_error(file_path, e)


class JSONExtractor(BaseExtractor):
    """Extractor for JSON files."""
    
    def extract(self, file_path: Path) -> str:
        """Extract text from JSON files."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                return json.dumps(data, indent=2, ensure_ascii=False)
        except Exception as e:
            return self._handle_extraction_error(file_path, e)


class ExcelExtractor(BaseExtractor):
    """Extractor for Excel files."""
    
    def extract(self, file_path: Path) -> str:
        """Extract text from Excel files."""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path)
            content_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content_parts.append(f"Sheet: {sheet_name}")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = '\t'.join([
                        str(cell) if cell is not None else '' 
                        for cell in row
                    ])
                    if row_text.strip():  # Only add non-empty rows
                        content_parts.append(row_text)
                
                content_parts.append("")  # Empty line between sheets
            
            return '\n'.join(content_parts).strip()
        except ImportError:
            return self._handle_import_error("Excel", "pip install openpyxl")
        except Exception as e:
            return self._handle_extraction_error(file_path, e)
